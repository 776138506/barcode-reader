"""R1 导出增强测试：两段式渲染、向后兼容、XLSX/JSON、过滤、剪贴板。"""
import json
import os
import sys
from pathlib import Path

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import pytest  # noqa: E402
from PySide6.QtCore import QSettings  # noqa: E402
from PySide6.QtWidgets import QApplication  # noqa: E402

from exporter import (ExportFilter, ExportRecord, apply_filter, export_json,  # noqa: E402
                      export_txt, export_xlsx, render_two_stage)
from ui import main_window as mw  # noqa: E402
from profiles import ProfileStore  # noqa: E402
from templates import TemplateStore  # noqa: E402

IMG_DIR = Path(__file__).resolve().parent / "images"

RECS = [ExportRecord("a.png", "QRCode", "abc"),
        ExportRecord("a.png", "QRCode", "def"),
        ExportRecord("b.png", "Code128", "ghi-123")]


@pytest.fixture(scope="module")
def qapp():
    app = QApplication.instance() or QApplication([])
    yield app


# ---------- 两段式渲染 ----------

def test_two_stage_group_image_tuple():
    """验收用例：行模板 {content} + 连接符 ',' + 外模板 {'{items}'} + 按图片分组。"""
    out = render_two_stage(RECS, "{content}", "','", "{'{items}'}", "image")
    assert out == "{'abc','def'}\n{'ghi-123'}"


def test_two_stage_group_global_and_sql():
    out = render_two_stage(RECS, "{content}", "','", "{'{items}'}", "global")
    assert out == "{'abc','def','ghi-123'}"
    sql = render_two_stage(RECS, "{content}", "','", "IN ('{items}')", "global")
    assert sql == "IN ('abc','def','ghi-123')"


def test_two_stage_none_backward_compatible(tmp_path):
    """外模板 {items} + 不分组 = 旧逐行行为；export_txt 默认参数输出不变。"""
    out = render_two_stage(RECS, "{content}", "\n", "{items}", "none")
    assert out == "abc\ndef\nghi-123"
    p = tmp_path / "out.txt"
    export_txt(RECS, str(p), "{index},{filename},{type},{content}")
    assert p.read_text(encoding="utf-8").splitlines() == [
        "1,a.png,QRCode,abc", "2,a.png,QRCode,def", "3,b.png,Code128,ghi-123"]


def test_two_stage_outer_literal_braces():
    """外模板除 {items} 外全按字面输出，单花括号不报格式错误。"""
    out = render_two_stage(RECS[:1], "{content}", ",", "[{{'{items}'}}]", "none")
    assert out == "[{{'abc'}}]"


# ---------- XLSX / JSON ----------

def test_export_xlsx(tmp_path):
    from openpyxl import load_workbook
    p = tmp_path / "out.xlsx"
    export_xlsx(RECS, str(p), "{index},{filename},{type},{content}",
                header="序号,文件名,码制,内容")
    ws = load_workbook(p).active
    rows = list(ws.iter_rows(values_only=True))
    assert rows[0] == ("序号", "文件名", "码制", "内容")
    assert rows[1] == ("1", "a.png", "QRCode", "abc")
    assert rows[3][3] == "ghi-123"


def test_export_json_grouping(tmp_path):
    p = tmp_path / "out.json"
    export_json(RECS, str(p), "none")
    flat = json.loads(p.read_text(encoding="utf-8"))
    assert len(flat) == 3 and flat[0]["content"] == "abc"
    assert set(flat[0]) == {"index", "filename", "type", "content", "count", "suspect"}

    export_json(RECS, str(p), "image")
    grouped = json.loads(p.read_text(encoding="utf-8"))
    assert len(grouped) == 2
    assert grouped[0]["group"] == "a.png"
    assert [i["content"] for i in grouped[0]["items"]] == ["abc", "def"]

    export_json(RECS, str(p), "global")
    agg = json.loads(p.read_text(encoding="utf-8"))
    assert len(agg) == 1 and agg[0]["group"] is None and len(agg[0]["items"]) == 3


# ---------- 导出过滤 ----------

def test_filter_individual_and_combined():
    assert apply_filter(RECS, ExportFilter(types=["QRCode"])) == RECS[:2]
    assert apply_filter(RECS, ExportFilter(min_len=4)) == RECS[2:]
    assert apply_filter(RECS, ExportFilter(max_len=3)) == RECS[:2]
    assert apply_filter(RECS, ExportFilter(prefix="gh")) == RECS[2:]
    assert apply_filter(RECS, ExportFilter(regex=r"^[a-z]{3}$")) == RECS[:2]
    # 组合：QRCode 且长度<=3 且前缀 a → 只剩 abc
    combo = ExportFilter(types=["QRCode"], max_len=3, prefix="a")
    assert apply_filter(RECS, combo) == [RECS[0]]
    combo2 = ExportFilter(types=["QRCode"], max_len=3, prefix="d")
    assert apply_filter(RECS, combo2) == [RECS[1]]
    # None/空 = 不限制
    assert apply_filter(RECS, ExportFilter()) == RECS
    assert apply_filter(RECS, None) == RECS


def test_filter_invalid_regex_raises():
    with pytest.raises(ValueError):
        apply_filter(RECS, ExportFilter(regex="(["))


# ---------- 剪贴板与持久化 ----------

def test_clipboard_rendered(qapp, tmp_path):
    settings = QSettings(str(tmp_path / "s.ini"), QSettings.IniFormat)
    win = mw.MainWindow(settings=settings, history_db=tmp_path / "h.db", **_stores(tmp_path))
    win.add_paths([str(IMG_DIR / "qr_hello.png"), str(IMG_DIR / "code128_a.png")])
    win._pool.waitForDone(30000)
    qapp.processEvents()

    win.template_edit.setText("{content}")
    win._joiner = "','"
    win._outer = "{'{items}'}"
    win._group_by = "global"
    win.copy_rendered_to_clipboard()
    text = QApplication.clipboard().text()
    assert text == "{'https://example.com/qr-hello','CODE128-ABC-12345'}"
    assert "2 条" in win.statusBar().currentMessage()

    # 过滤后复制：只留 Code128
    win._filter = ExportFilter(types=["Code128"])
    win.copy_rendered_to_clipboard()
    assert QApplication.clipboard().text() == "{'CODE128-ABC-12345'}"
    win.close()


def test_export_settings_persist(qapp, tmp_path):
    settings = QSettings(str(tmp_path / "s.ini"), QSettings.IniFormat)
    win = mw.MainWindow(settings=settings, history_db=tmp_path / "h.db", **_stores(tmp_path))
    win._joiner = "','"
    win._outer = "{'{items}'}"
    win._group_by = "image"
    win._filter = ExportFilter(types=["QRCode"], min_len=2, max_len=99,
                               prefix="ab", regex="x+")
    win._save_settings()
    win.close()

    win2 = mw.MainWindow(settings=QSettings(str(tmp_path / "s.ini"), QSettings.IniFormat),
                         history_db=tmp_path / "h2.db", **_stores(tmp_path))
    assert win2._joiner == "','"
    assert win2._outer == "{'{items}'}"
    assert win2._group_by == "image"
    assert win2._filter.types == ["QRCode"]
    assert win2._filter.min_len == 2 and win2._filter.max_len == 99
    assert win2._filter.prefix == "ab" and win2._filter.regex == "x+"
    win2.close()


def _stores(tmp_path):
    """tmp store 注入：测试不碰真实用户数据目录（AGENTS 纪律）。"""
    return {"profile_store": ProfileStore(tmp_path / "profiles.json"),
            "template_store": TemplateStore(tmp_path / "templates.json")}
